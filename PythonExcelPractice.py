# Read ProduceReport
# Write all contents of this file to a Second Sheet in the current workbook
# Display Grand Total and Average of AMT SOLD and TOTAL

import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title='Second Sheet')

#

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row

write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost per Round'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'

write_row = 2

write_colA = 1
write_colB = 2
write_colC = 3
write_colD = 4

for currentrow in read_ws.iter_rows(min_row=2, max_row=maxR, max_col = maxC):
    
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = float(currentrow[2].value)
    total = float(currentrow[3].value)

    write_sheet.cell(write_row, write_colA).value = name
    write_sheet.cell(write_row, write_colB).value = cost
    write_sheet.cell(write_row, write_colC).value = amt_sold
    write_sheet.cell(write_row, write_colD).value = total

    write_row += 1

summary_row = write_row + 1

write_sheet['B' + str(summary_row)] = 'Total'
write_sheet['B' + str(summary_row)].font = Font(size=16, bold=True)

write_sheet['C' + str(summary_row)] = '=SUM(C2:C' + str(write_row) + ')'
write_sheet['D' + str(summary_row)] = '=SUM(D2:D' + str(write_row) + ')'

summary_row = 1

write_sheet['B' + str(summary_row)] = 'Average'
write_sheet['B' + str(summary_row)].font = Font(size=16, bold=True)

write_sheet['C' + str(summary_row)] = '=Average(C2:C' + str(write_row) + ')'
write_sheet['D' + str(summary_row)] = '=Average(D2:D' + str(write_row) + ')'

write_sheet.column_dimensions['A'] = 16
write_sheet.column_dimensions['B'] = 15
write_sheet.column_dimensions['C'] = 15
write_sheet.column_dimensions['D'] = 15

for cell in write_sheet["C:C"]:
    cell.number_format = '#,##0'

for cell in write_sheet["D:D"]:
    cell.number_format = u'#$ "#,##0.00'

wb.save('PythontoExcel.xlsx')

