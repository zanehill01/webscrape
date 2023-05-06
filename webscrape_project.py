
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup

import openpyxl as xl
from openpyxl.styles import Font

import keys
from twilio.rest import Client

# twilio setup

client = Client(keys.account_sid, keys.auth_token)
TwilioNumber = "+18775254432"
mycellphone = "+15127881279"

# webscrape setup

webpage = 'https://www.webull.com/quote/crypto'
page = urlopen(webpage)
soup = BeautifulSoup(page, 'html.parser')
title = soup.title
data = soup.findAll('div', class_ = 'table-cell')

# excel setup

wb = xl.Workbook()
ws = wb.active
ws.title = ('crypto info')
write_sheet = wb['crypto info']

# font setup

title_font = Font(name='Times New Roman', size=18, italic=False, bold=True)
pos_font = Font(name='Times New Roman', size = 14, color='00FF00')
neg_font = Font(name='Times New Roman', size = 14, color='FF0000')
reg_font = Font(name='Times New Roman', size = 14)

# fill excel headers

ws['A1'] = " # "
ws['A1'].font = title_font

ws['B1'] = "Crypto Currency"
ws['B1'].font = title_font

ws['C1'] = "Symbol"
ws['C1'].font = title_font

ws['D1'] = "Current Price"
ws['D1'].font = title_font

ws['E1'] = "Percent Change (24 hr)"
ws['E1'].font = title_font

ws['F1'] = "New Price"
ws['F1'].font = title_font

# establish dimensions for headers

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 20

# counter variables

counter = 1
excel_counter = 1
prices = []

# loop, validate data and compile excel information

for x in range(1, 6):

    # create name and symbol, remove symbol from name

    name = data[counter].text

    symbol = name[1:name.index('USD')+3]
    name = name.replace(symbol, '')[1:]

    # error catching for missing values

    try:
        current_price = float(data[counter+1].text.replace(',',''))
    except:
        current_price = data[counter+2].text.replace(',','')

    # find percent change

    percent_change = data[counter+2].text
    new_change = round((float(percent_change.replace('%','')) / 100) * current_price, 4)

    # question if percent_change is BTC or ETH and < or > 5, create text variable which sends text

    if symbol == 'BTCUSD':
        if new_change >= -5 and new_change <= 5:
            twilio_text = Client.messages.create(to=mycellphone, from_=TwilioNumber, body='BitCoin (BTC) has changed +/- within 5 $')
    if symbol == 'ETHUSD':
        if new_change >= -5 and new_change <= 5:
            twilio_text = Client.messages.create(to=mycellphone, from_=TwilioNumber, body='Etherium (ETC) has changed +/- within 5 $')

    # fill excel document with data

    if counter < 51:

        ws['A' + str(excel_counter+1)] = excel_counter
        ws['B' + str(excel_counter+1)] = name
        ws['C' + str(excel_counter+1)] = symbol

        ws['A' + str(excel_counter+1)].font = reg_font
        ws['B' + str(excel_counter+1)].font = reg_font
        ws['C' + str(excel_counter+1)].font = reg_font     

        if '-' in percent_change: 

            ws['D' + str(excel_counter+1)] = '$' + str(round(current_price, 4))
            ws['E' + str(excel_counter+1)] = percent_change
            ws['F' + str(excel_counter+1)] = '$' + str(round(new_change + current_price, 4))

            ws['D' + str(excel_counter+1)].font = neg_font
            ws['E' + str(excel_counter+1)].font = neg_font
            ws['F' + str(excel_counter+1)].font = neg_font

        else:

            ws['D' + str(excel_counter+1)] = '$' + str(round(current_price, 4))
            ws['E' + str(excel_counter+1)] = percent_change
            ws['F' + str(excel_counter+1)] = '$' + str(round(new_change + current_price, 4))

            ws['D' + str(excel_counter+1)].font = pos_font
            ws['E' + str(excel_counter+1)].font = pos_font
            ws['F' + str(excel_counter+1)].font = pos_font

        excel_counter += 1
        counter += 10

#

wb.save('cryptoinfo.xlsx')
print('Done')