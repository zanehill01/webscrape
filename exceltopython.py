import openpyxl as xl

wb = xl.load_workbook('example.xlsx')

sn = wb.sheetnames

print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1)

print(cellA1.value)
print(type(cellA1.value))

print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

for i in range(1, sheet1.max_row + 1):
    print(sheet1.cell(i, 2).value)

# convert letters to numbers

print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(900))

for currentrow in sheet1['A1':'C3']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)

for currentrow in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, max_col=sheet1.max_column):
    print(currentrow)