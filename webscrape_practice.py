
import random
from urllib.request import urlopen
from bs4 import BeautifulSoup
from urllib.request import urlopen, Request
import openpyxl as xl

webpage = 'https://www.boxofficemojo.com/year/2023/?ref_=bo_yl_table_2'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)
webpage = urlopen(req).read()
soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title.text)

all_tables = soup.findAll('table')

movie_table = all_tables[0]

movie_rows = movie_table.findAll('tr')

wb = xl.Workbook()
ws = wb.active

ws.title = 'BoxOfficeReport'

ws['A1'] = 'Number'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Total Gross'
ws['F1'] = '% of Total Gross'

for x in range(1, 6):
    td = movie_rows[x].findAll('td')

    no = td[0].text
    title = td[1].text
    gross = int(td[5].text.replace(',','').replace('$',''))
    total_gross = int(td[7].text.replace(',','').replace('$',''))
    release_date = td[8].text

    percent_gross = round((gross/total_gross)*100, 2)

    ws['A' + str(x+1)] = no
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = release_date
    ws['D' + str(x+1)] = gross
    ws['E' + str(x+1)] = total_gross
    ws['F' + str(x+1)] = str(percent_gross) + '%'

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 26

wb.save('BoxOfficeReport.xlsx')