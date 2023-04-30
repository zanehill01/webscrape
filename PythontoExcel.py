import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active()

ws.title = 'FirstSheet'

wb.create_sheet(index=1,title='SecondSheet')

ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman', size=24, italics=False, bold=True)

myfont = Font(name='Times New Roman', size=24, italics=False, bold=True)

ws.merge_cells('A1:B1')
ws.unmerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

wb.save('PythontoExcel.xlsx')